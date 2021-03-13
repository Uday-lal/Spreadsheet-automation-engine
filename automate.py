"""
--------------------------------------------------------->
Copyrights (c) to UR's tech.ltd 2021. All rights reserved
Author: Uday lal
company: UR's tech.ltd
--------------------------------------------------------->

~~~~~~~~~~~~~~~~~~~~~~~
Making the automation engine.
"""
from openpyxl import load_workbook


def splitting_algorithm(wb_data):
    """
    Splitting the wb data into several parts
    :param wb_data: input wb data
    :return: dict
    """
    sheets = wb_data["sheets"]
    intro_part = []
    return_data = {}
    for sheet in sheets:
        current_sheet = wb_data[sheet]
        rows = current_sheet["rows"]

        for i in range(len(rows)):  # Checking for the heading
            current_row = rows[i]
            validation = Validation(data=current_row)
            if "heading" not in return_data:
                if validation.is_heading():
                    return_data["heading"] = current_row
                    break

        for i in range(len(rows)):  # Checking for the intro part
            current_row = rows[i]
            validation = Validation(data=current_row)
            if validation.is_intro_part():
                intro_part.append(current_row)
                return_data["intro_part"] = intro_part

        if "intro_part" not in return_data:
            return_data["intro_part"] = None

    return return_data


class FileTypeException(Exception):
    def __init__(self, message):
        self.error_message = message


class Validation:
    def __init__(self, data):
        self.data = data

    def is_heading(self):
        """
        Validating the data as the heading
        :return: bool
        """
        none_times = self.times(data_array=self.data, selected_item="")
        data_length = len(self.data)

        if none_times is None:
            none_times = 0

        not_none_times = data_length - none_times
        if not_none_times != 1 and not_none_times != 0:
            return True
        return False

    def is_intro_part(self):
        """
        Validating the data as the intro part
        :return: bool
        """
        none_times = self.times(data_array=self.data, selected_item="")
        data_length = len(self.data)

        if none_times is None:
            none_times = 0

        not_none_times = data_length - none_times

        if not_none_times == 1:
            return True
        return False

    @staticmethod
    def times(data_array, selected_item):
        """
        Collecting the index of the selected element
        :param data_array: Input data array
        :param selected_item: Element needs to search
        :return: int
        """
        itemCount = 0
        if selected_item not in data_array:
            return None

        for item in data_array:
            if item != selected_item:
                continue
            else:
                itemCount += 1

        return itemCount


class Automate:
    def __init__(self, filename):
        self.filename = str(filename)
        if not self.filename.endswith(".xlsx"):
            raise FileTypeException(
                f"Excepting .xlsx file got {self.filename[self.filename.index('.'):len(self.filename)]}")

    def get_workbook_data(self):
        """
        Fetch the workbook data
        :return: dict
        """
        wb = load_workbook(self.filename)
        sheets = wb.sheetnames
        wb_data = {"sheets": sheets}

        for sheet in sheets:
            current_sheet = wb[sheet]
            max_cols = current_sheet.max_column
            max_rows = current_sheet.max_row
            rows_data = []
            sheet_data = {}
            rows = []

            for row in range(1, max_rows + 1):
                for col in range(1, max_cols + 1):
                    cell_data = current_sheet.cell(row, col).value
                    if cell_data is None:
                        cell_data = ""

                    rows_data.append(cell_data)

                cols_data_copy = rows_data.copy()
                rows.append(cols_data_copy)
                sheet_data["rows"] = rows
                sheet_data["max_cols"] = max_cols
                sheet_data["max_row"] = max_rows
                rows_data.clear()

            wb_data[sheet] = sheet_data

        return wb_data
