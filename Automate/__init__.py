"""
~~~~~~~~~~~~~~~~~~~~~~~
Starting the main automation class.
"""
from openpyxl import load_workbook
from mof_library.get_workbook_data import GetWbData
from mof_library.save_workbook_data import SaveWorkbookData


def splitting_algorithm(wb_data):
    """
    Splitting the wb data into several parts
    :param wb_data: input wb data
    :return: dict
    """
    sheets = wb_data["sheets"]
    intro_part = []
    return_data = {}
    for sheet in sheets:
        current_sheet = wb_data[sheet]
        rows = current_sheet["rows"]

        for i in range(len(rows)):  # Checking for the heading
            current_row = rows[i]
            validation = Validation(data=current_row)
            if "heading" not in return_data:
                if validation.is_heading():
                    return_data["heading"] = current_row
                    break

        for i in range(len(rows)):  # Checking for the intro part
            current_row = rows[i]
            validation = Validation(data=current_row)
            if validation.is_intro_part():
                intro_part.append(current_row)
                return_data["intro_part"] = intro_part

        if "intro_part" not in return_data:
            return_data["intro_part"] = None

    return return_data


class FileTypeException(Exception):
    def __init__(self, message):
        self.error_message = message


class Validation:
    def __init__(self, data):
        self.data = data

    def is_heading(self):
        """
        Validating the data as the heading
        :return: bool
        """
        none_times = self.times(data_array=self.data, selected_item="")
        data_length = len(self.data)

        if none_times is None:
            none_times = 0

        not_none_times = data_length - none_times
        if not_none_times != 1 and not_none_times != 0:
            return True
        return False

    def is_intro_part(self):
        """
        Validating the data as the intro part
        :return: bool
        """
        none_times = self.times(data_array=self.data, selected_item="")
        data_length = len(self.data)

        if none_times is None:
            none_times = 0

        not_none_times = data_length - none_times

        if not_none_times == 1:
            return True
        return False

    @staticmethod
    def times(data_array, selected_item):
        """
        Collecting the index of the selected element
        :param data_array: Input data array
        :param selected_item: Element needs to search
        :return: int
        """
        itemCount = 0
        if selected_item not in data_array:
            return None

        for item in data_array:
            if item != selected_item:
                continue
            else:
                itemCount += 1

        return itemCount


class Automate:
    def __init__(self, filename=None):
        self.filename = str(filename)
        if not self.filename.endswith(".xlsx") and filename is not None:
            raise FileTypeException(
                f"Excepting .xlsx file got {self.filename[self.filename.index('.'):len(self.filename)]}")

    def get_workbook_data(self):
        """
        Fetch the workbook data
        :return: dict
        """
        wb = load_workbook(self.filename)
        gwd = GetWbData(wb_obj=wb, filename=self.filename)
        wb_data = gwd.get_data()
        return wb_data

    @staticmethod
    def sort(data):
        return sorted(list(data))

    def reverse(self, data):
        sorted_data = self.sort(data)
        sorted_data.reverse()
        return sorted_data  # Return reverse sorted data

    def delete(self, wb_data, selected_index):
        self.wb_data = wb_data
        _data = self.get_data()
        while True:
            try:
                row_data = next(_data)
                del row_data[selected_index]
            except StopIteration:
                break

    def get_data(self):
        for row_data in self.wb_data:
            yield row_data

    @staticmethod
    def save_wb(file_path, data, updated_path):
        wb = load_workbook(file_path)
        save_wb = SaveWorkbookData(wb_obj=wb, wb_data=data, updated_path=updated_path)
        save_wb.save_data()
